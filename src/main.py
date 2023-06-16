from tkinter import *
import customtkinter as ctk
import webbrowser
import subprocess
import os
import openpyxl
import pathlib
from openpyxl import Workbook
from tkinter.filedialog import asksaveasfilename, askopenfilename
from PIL import Image, ImageTk
from tkinter import messagebox
from tkinter.ttk import Combobox
from stegano import lsb
from calculator import Calculator
from signup_and_login_form import SignUp
from definition import DictionaryDefinition
from mytranslator import PythonTranslator
from notepad import NotePad
from writing_board import WritingBoard
from yt_video_downloader import VideoDownloader
from clock import Clock
import settings 
import report_page 
import keyboard_shortcut
from grade_calculator import Grade
from encrypting import EncryptDecrypt
from get_apartment import Apartment
from excel_viewer import ExcelViewer
import mcq



root = Tk()
root.geometry('850x830+200+50')
root.minsize(300, 300)
root.title('Menu')
root.configure(background='#d4c1e6')

# global variables
file = None
username = '@Unknown'



############# all command's execute here #############

def bangladesh():
    webbrowser.open('https://en.wikipedia.org/wiki/Bangladesh')
    
    
def information():
    root.lift()
    messagebox.showinfo('Information', "This is a project of HPI skills competition. It will be updateable!\n@MD_Shakib_Ahmed all right and reserved.", parent=root)


def share():
    link = ''
    

def open_youtube():
    webbrowser.open('youtube.com')


def open_facebook():
    webbrowser.open('facebook.com')
    


def open_python():
    webbrowser.open('https://docs.python.org/3/')
    
    
    
def open_sof():
    webbrowser.open('stackoverflow.com')
    
    
def open_github():
    webbrowser.open('github.com')



def open_chatgpt():
    webbrowser.open('chat.openai.com')



def open_account():
    pass



def quiz_game():
    root = Toplevel()
    quiz = mcq.MCQquiz(root)
    root.mainloop()



def translate_anything():
    trans = PythonTranslator()
    trans.run()



def open_notepad():
    npad = NotePad()



def open_calculator():
    window = Toplevel()
    calc = Calculator(window)
    window.mainloop()


def open_downloadar():
    d1 = VideoDownloader()
    d1.run()



def open_writingboard():
    w1 = WritingBoard()
    w1.run()
    



def singup_form():
    global username
    
    signup1 = SignUp()
    username = signup1.username # get username
    
    
def excel_file():
    excel = ExcelViewer()
    excel.excel_view()
    



################################# Student Registration System strats ############################
def student_registration():
    
    root = Toplevel()
    root.title('Registration Form')
    root.geometry('1400x750')
    # values of combobox
    technologys = ['Computer Science & Technology (85)',
                'Architecture Technology (61)',
                'Civil Technology (64)',
                'Electronics Technology (68)',
                'Mechanical Technology (70)',
                'Power Technology (71)',
                'Civil (Wood) Technology (65)'
                ]

    


    file = pathlib.Path('student_data.xlsx')
    if file.exists():
        pass

    else:
        file = Workbook()
        sheet = file.active
        sheet['A1'] = 'Registration Number'
        sheet['B1'] = "Student's Name"
        sheet['C1'] = 'Father\'s Name'
        sheet['D1'] = 'Mohter\'s Name'
        sheet['E1'] = 'Roll No.'
        sheet['F1'] = 'Gender'
        sheet['G1'] = 'Technology Name & Code'
        sheet['H1'] = 'Date of Birth'
        sheet['I1'] = 'Institute Name & Code'
        sheet['J1'] = 'Post Office'
        sheet['K1'] = 'Upazilla/Thana'
        sheet['L1'] = 'District'
        sheet['M1'] = 'Session'

        file.save('student_data.xlsx')


    # all commands execute here...
    def upload_img():
        global file_name, img
        root.lift()
        file_name = askopenfilename(initialdir=os.getcwd(),
                                            title='Select Image File',
                                            filetypes=(
                                                ('JPG File', '*.jpg'),
                                                ('PNG File', '*.png'),
                                                ('All Files', '*.*')
        ), parent=root)

        img = Image.open(file_name)
        img_resize = img.resize((200, 200))
        img_ = ImageTk.PhotoImage(image=img_resize)

        img_lbl.config(image=img_)
        img_lbl.image = img_


    def save():
        regi_value = regi.get()  
        name_value = name.get()
        f_name_value = f_name.get()
        m_name_value = m_name.get()
        roll_value = roll.get()   
        gender_value = genderVar.get()
        tech_value = tech_name.get()
        dob_value = dob.get()
        ins_value = ins_nameVar.get()
        p_office_value = post_officeVar.get()
        thana_value = thanaVar.get()
        district_value = districtVar.get()
        session_value = sessionVar.get()
        
        
        if regi_value != '':
            try:
                regi_value = int(regi_value)
            except:
                root.lift()
                messagebox.showerror('Error', 'Registration Number Cannot be string.', parent=root)
        
        
        if roll_value != '':
            try:
                roll_value = int(roll_value)
            except:
                root.lift()
                messagebox.showwarning('Roll', 'Roll No. cannot be string.', parent=root)
        
        
        if regi_value == '':
            root.lift()
            messagebox.showerror('Registration', 'Registration Number Cannot be Empty!', parent=root)
            
        elif name_value == '':
            root.lift()
            messagebox.showerror('Name', 'Student Name Cannot be Empty!', parent=root)
            
        elif tech_value == '':
            root.lift()
            messagebox.showerror('Technology', 'Technology Cannot be Empty', parent=root)
            
        elif ins_value == '':
            root.lift()
            messagebox.showerror('Institute', 'Institute Cannot be Empty!', parent=root)
            
        else:
            file = openpyxl.load_workbook('student_data.xlsx')
            sheet = file.active
            
            sheet.cell(column=1, row=sheet.max_row+1, value=regi_value)
            sheet.cell(column=2, row=sheet.max_row, value=name_value)
            sheet.cell(column=3, row=sheet.max_row, value=f_name_value)
            sheet.cell(column=4, row=sheet.max_row, value=m_name_value)
            sheet.cell(column=5, row=sheet.max_row, value=roll_value)
            sheet.cell(column=6, row=sheet.max_row, value=gender_value)
            sheet.cell(column=7, row=sheet.max_row, value=tech_value)
            sheet.cell(column=8, row=sheet.max_row, value=dob_value)
            sheet.cell(column=9, row=sheet.max_row, value=ins_value)
            sheet.cell(column=10, row=sheet.max_row, value=p_office_value)
            sheet.cell(column=11, row=sheet.max_row, value=thana_value)
            sheet.cell(column=12, row=sheet.max_row, value=district_value)
            sheet.cell(column=13, row=sheet.max_row, value=session_value)
            
            file.save(r'student_data.xlsx')
            
            try:
                img.save('Student Images/' + str(regi_value) + '.jpg')
            except:
                root.lift()
                messagebox.showinfo('Info', "Profile picture is not available or configured.", parent=root)
        
            root.lift()
            messagebox.showinfo('Info', "Successfully data entered!", parent=root)
            
            # call the reset function
            reset()
        


    def reset():
        global img

        regi.delete(0, 'end')
        name.delete(0, 'end')
        f_name.delete(0, 'end')
        m_name.delete(0, 'end')
        roll.delete(0, 'end')
        dob.delete(0, 'end')
        regi_search.delete(0, 'end')

        save_btn.config(state='normal')

        user_img = Image.open('assets/user.png')
        user_img = user_img.resize((200, 200))
        user_img = ImageTk.PhotoImage(image=user_img)
        img_lbl.config(image=user_img)
        img_lbl.image = user_img

        img = ''



    ### search ###
    def search():
        search_text = regi_search.get()

        reset()  # to reset everything
        save_btn.config(state='disable')  # after clicking search, save button will be disabled

        file = openpyxl.load_workbook('student_data.xlsx')
        sheet = file.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == int(search_text):
                regi_value = row[0]
                name_value = row[1]
                f_name_value = row[2]
                m_name_value = row[3]
                roll_value = row[4]
                gender_value = row[5]
                tech_value = row[6]
                dob_value = row[7]
                ins_value = row[8]
                p_office_value = row[9]
                thana_value = row[10]
                district_value = row[11]
                session_value = row[12]

                regi.delete(0, 'end')
                regi.insert(0, regi_value)
                name.delete(0, 'end')
                name.insert(0, name_value)
                f_name.delete(0, 'end')
                f_name.insert(0, f_name_value)
                m_name.delete(0, 'end')
                m_name.insert(0, m_name_value)
                roll.delete(0, 'end')
                roll.insert(0, roll_value)
                genderVar.set(gender_value)
                tech_name.set(tech_value)
                dob.delete(0, 'end')
                dob.insert(0, dob_value)
                ins_nameVar.set(ins_value)
                post_officeVar.set(p_office_value)
                thanaVar.set(thana_value)
                districtVar.set(district_value)
                sessionVar.set(session_value)

                try:
                    img = Image.open(f'Student Images/{regi_value}.jpg')
                    img_resize = img.resize((200, 200))
                    img_ = ImageTk.PhotoImage(image=img_resize)

                    img_lbl.config(image=img_)
                    img_lbl.image = img_
                except:
                    root.lift()
                    messagebox.showinfo('Info', "Profile picture is not available or configured.", parent=root)

                break
            
            else:
                pass

        file.close()
        
        
        
    def update():
        regi_value = regi.get()
        try:
            regi_value = int(regi_value)
        except:
            root.lift()
            messagebox.showerror('Error', 'Registration Number Cannot be a string.', parent=root)

        name_value = name.get()
        f_name_value = f_name.get()
        m_name_value = m_name.get()
        roll_value = roll.get()
        try:
            roll_value = int(roll_value)
        except:
            root.lift()
            messagebox.showwarning('Roll', 'Roll No. cannot be a string.', parent=root)

        gender_value = genderVar.get()
        tech_value = tech_name.get()
        dob_value = dob.get()
        ins_value = ins_nameVar.get()
        p_office_value = post_officeVar.get()
        thana_value = thanaVar.get()
        district_value = districtVar.get()
        session_value = sessionVar.get()

        if regi_value == '':
            root.lift()
            messagebox.showerror('Registration', 'Registration Number Cannot be Empty!', parent=root)
        elif name_value == '':
            root.lift()
            messagebox.showerror('Name', 'Student Name Cannot be Empty!', parent=root)
        elif tech_value == '':
            root.lift()
            messagebox.showerror('Technology', 'Technology Cannot be Empty', parent=root)
        elif ins_value == '':
            root.lift()
            messagebox.showerror('Institute', 'Institute Cannot be Empty!', parent=root)
        else:
            file = openpyxl.load_workbook('student_data.xlsx')
            sheet = file.active

            found = False
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == regi_value:
                    found = True
                    sheet.cell(column=2, row=sheet.max_row, value=name_value)
                    sheet.cell(column=3, row=sheet.max_row, value=f_name_value)
                    sheet.cell(column=4, row=sheet.max_row, value=m_name_value)
                    sheet.cell(column=5, row=sheet.max_row, value=roll_value)
                    sheet.cell(column=6, row=sheet.max_row, value=gender_value)
                    sheet.cell(column=7, row=sheet.max_row, value=tech_value)
                    sheet.cell(column=8, row=sheet.max_row, value=dob_value)
                    sheet.cell(column=9, row=sheet.max_row, value=ins_value)
                    sheet.cell(column=10, row=sheet.max_row, value=p_office_value)
                    sheet.cell(column=11, row=sheet.max_row, value=thana_value)
                    sheet.cell(column=12, row=sheet.max_row, value=district_value)
                    sheet.cell(column=13, row=sheet.max_row, value=session_value)

                    try:
                        img.save('Student Images/' + str(regi_value) + '.jpg')
                    except:
                        root.lift()
                        messagebox.showinfo('Info', "Profile picture is not available or configured.", parent=root)
                    
                    root.lift()
                    messagebox.showinfo('Info', "Successfully updated the data!", parent=root)
                    break

            file.save(r'student_data.xlsx')

            if not found:
                root.lift()
                messagebox.showinfo('Info', 'No matching registration number found.', parent=root)

            reset()





    # title frame
    title_f = Frame(root, bg='orange',
                    height=100, width=1400,
                    borderwidth=1,
                    )
    title_f.place(x=0, y=0)

    # title
    Label(title_f, text='Registration System Diploma in Engineering',
        font=('Arial', 26), bg='orange').place(x=10, y=20)

    Label(title_f, text='[Curriculum Code: 15]', bg='orange', ).place(x=250, y=70)

    # search box
    regi_search = Entry(title_f, width=25,
                        font=('Arial', 18),
                        borderwidth=0
                        )
    regi_search.place(x=900, y=30)

    # search button
    Button(title_f, text='Search', bg='white', borderwidth=0, relief='groove',
        font=('Arial', 13), command=search).place(x=1229, y=30)

    # regi no.
    Label(root, text='Registration Number :', font=(
        'Arial', 12, 'bold')).place(x=10, y=124)
    regi = Entry(root, font=("Arial", 16), borderwidth=0,)
    regi.place(x=200, y=120)

    # session
    sessionVar = StringVar(root, value='2021-22')

    Label(root, text='Session :', font=('Arial', 12, 'bold')).place(x=700, y=124)
    session = Entry(root, font=("Arial", 16),
                    borderwidth=0, textvariable=sessionVar)
    session.place(x=810, y=120)


    # data frames #
    first_f = Frame(root, borderwidth=1, bg='white', height=250, width=1050)
    second_f = Frame(root, borderwidth=1, bg='white', height=250, width=1050)
    third_f = Frame(root, borderwidth=1, bg='white', height=570, width=280)

    first_f.place(x=20, y=180)
    second_f.place(x=20, y=440)
    third_f.place(x=1100, y=120)


    # data in first frame
    # labels
    Label(first_f, text='Name of the Student :',
        bg='white', font=('Arial', 12)).place(x=10, y=20)
    Label(first_f, text="Father's Name :", bg='white',
        font=('Arial', 12)).place(x=10, y=70)
    Label(first_f, text='Mother\'s Name :', bg='white',
        font=('Arial', 12)).place(x=10, y=120)
    Label(first_f, text='Technology Name & Code :',
        bg='white', font=('Arial', 12)).place(x=10, y=170)

    # entries
    name = Entry(first_f, font=("Arial", 16), borderwidth=1)
    f_name = Entry(first_f, font=("Arial", 16), borderwidth=1)
    m_name = Entry(first_f, font=("Arial", 16), borderwidth=1)
    tech_name = Combobox(first_f, font=("Arial", 14),
                        values=technologys, cursor='hand2', width=24)

    name.place(x=220, y=17)
    f_name.place(x=220, y=67)
    m_name.place(x=220, y=117)
    tech_name.place(x=220, y=167)

    # gender
    genderVar = StringVar(first_f, value='Male')
    Label(first_f, text='Gender :', font=(
        'Arial', 12), bg='white').place(x=600, y=20)
    gender = Entry(first_f, font=("Arial", 14), bg='white',
                borderwidth=1, textvariable=genderVar)
    gender.place(x=750, y=20)


    # roll number
    Label(first_f, text='Roll No. :', font=(
        'Arial', 12), bg='white').place(x=600, y=70)
    roll = Entry(first_f, font=("Arial", 14), bg='white', borderwidth=1)
    roll.place(x=750, y=70)

    # date of birth
    Label(first_f, text='Date of Birth :', font=(
        'Arial', 12), bg='white').place(x=600, y=120)
    dob = Entry(first_f, font=("Arial", 14), bg='white', borderwidth=1)
    dob.place(x=750, y=120)


    # data in second frame
    # labels
    Label(second_f, text='Institute Name & Code :',
        bg='white', font=('Arial', 12)).place(x=10, y=20)
    Label(second_f, text="Post Office :", bg='white',
        font=('Arial', 12)).place(x=10, y=70)
    Label(second_f, text='Upazilla/Thana :', bg='white',
        font=('Arial', 12)).place(x=10, y=120)
    Label(second_f, text='District :',
        bg='white', font=('Arial', 12)).place(x=10, y=170)

    # text variables
    ins_nameVar = StringVar(
        second_f, value='Habiganj Polytechnic Institute (63010)')
    post_officeVar = StringVar(second_f, value='Gopaya')
    thanaVar = StringVar(second_f, value='Habiganj Sadar')
    districtVar = StringVar(second_f, value='Habiganj')

    # entries
    ins_name = Entry(second_f, font=("Arial", 16), borderwidth=1,
                    textvariable=ins_nameVar, width=32)
    post_office = Entry(second_f, font=("Arial", 16),
                        borderwidth=1, textvariable=post_officeVar, width=32)
    thana = Entry(second_f, font=("Arial", 16), borderwidth=1,
                textvariable=thanaVar, width=32)
    district = Entry(second_f, font=("Arial", 16), borderwidth=1,
                    textvariable=districtVar, width=32)

    ins_name.place(x=220, y=17)
    post_office.place(x=220, y=67)
    thana.place(x=220, y=117)
    district.place(x=220, y=167)


    # data in third frame
    # img frame
    img_f = Frame(third_f, bg='white', height=240, width=240, highlightthickness=1)
    img_f.place(x=20, y=20)

    # image
    user_img = Image.open('assets/user.png')
    user_img = user_img.resize((200, 200))
    user_img = ImageTk.PhotoImage(image=user_img)
    img_lbl = Label(img_f, image=user_img)
    img_lbl.place(x=20, y=20)


    # buttons
    upload_btn = Button(third_f, text='Upload', height=3, width=20,
                        command=upload_img, bg='skyblue', bd=0)
    upload_btn.place(x=45, y=280)

    save_btn = Button(third_f, text='Save', height=3, width=20,
                    command=save, bg='#44dd99', bd=0)
    save_btn.place(x=45, y=350)

    reset_btn = Button(third_f, text='Reset', height=3, width=20,
                    command=reset, bg='orange', bd=0)
    reset_btn.place(x=45, y=420)

    Button(third_f, text='Update', height=3, width=20,
        command=update, bg='#44ffff', bd=0).place(x=45, y=490)
    
    
    
    
    ##################### Student Registration System Ends ##############################



def house_price():
    apartment = Apartment()
    apartment.find_apartment()



def report_issue():
    r1 = report_page.Report()



def open_calendar():
    c1 = Clock()
    c1.main()



def open_dictionary():
    root = Toplevel()
    dict1 = DictionaryDefinition(root)
    root.mainloop()



def grade_calc():
    grd = Grade()
    grd.grade_calculate()



def data_encryption():
    ed = EncryptDecrypt()
    ed.encrypt_decrypt()
    
    
    
    
def data_hide():
    root = Toplevel()
    root.title('Hide Your Data')
    root.geometry('900x540')
    root.configure(background='#307791')
    
    
    
    ################## Hide Data starts #########################
    
    
    ### all executable functions
    def open_image():
        global file
        root.lift()
        file = askopenfilename(initialdir=os.getcwd(),
                            title='Selcet Image Files',
                            filetypes=(
                                ('All Files', '*.*'),
                                ('PNG Files', '*.png'),
                                ('JPG Files', '*.jpg')
                            ), parent=root)
        
        
        img = Image.open(file)
        img_resize = img.resize((340, 340))
        img_ = ImageTk.PhotoImage(img_resize)
        
        img_lbl.configure(image=img_)
        img_lbl.image = img_
        


    def save_image():
        secret_msg.save(fp='hidden_data.png')


    def hidedata():
        global secret_msg
        msg = text_area.get(1.0, 'end')
        secret_msg = lsb.hide(str(file), message=msg)


    def showdata():
        real_msg = lsb.reveal(file)
        
        text_area.delete(1.0, 'end')
        text_area.insert('end', real_msg)





    # title label
    Label(root, text='Is Your Password Secured?',
        font=('Arial', 26),
        background='#307791'
        ).pack(padx=5, pady=10)



    # image frame
    img_frame = ctk.CTkFrame(root, 
                    height=350, 
                    width=420, 
                    fg_color='#153c4a',
                    )

    img_frame.place(x=20, y=80)

    # label for plot image 
    img_lbl = Label(img_frame, bg='#153c4a')
    img_lbl.place(x=40, y=10)



    # text fill frame
    text_frame = ctk.CTkFrame(root, 
                    height=350, 
                    width=420, 
                    fg_color='#153c4a',
                    )

    text_frame.place(x=450, y=80)

    text_area = ctk.CTkTextbox(text_frame,
                    height=348,
                    width=418,
                    font=('Arial', 25),
                    fg_color='#153c4a'
                    )
    text_area.place(x=1, y=1)


    ### buttons ###
    btn_frame = ctk.CTkFrame(root, width=430,
                        height=60,
                        fg_color='#153c4a',
                        border_color='white', 
                        border_width=1,
                        )
    btn_frame.place(x=10, y=440)


    # open image
    open_img = ctk.CTkButton(btn_frame, text='Open Image', 
                        width=200, height=40,
                        font=('Arial', 18),
                        corner_radius=10,
                        bg_color='#153c4a', 
                        hover_color='blue',
                        command=open_image
                        )
    open_img.place(x=5, y=10)


    # save image
    save_img = ctk.CTkButton(btn_frame, text='Save Image', 
                        width=200, height=40,
                        font=('Arial', 18),
                        corner_radius=10,
                        bg_color='#153c4a', 
                        hover_color='blue',
                        command=save_image
                        )
    save_img.place(x=220, y=10)


    # frame 2 #
    btn_frame2 = ctk.CTkFrame(root, width=430,
                        height=60,
                        fg_color='#153c4a',
                        border_color='white', 
                        border_width=1,
                        )
    btn_frame2.place(x=450, y=440)


    # Hide Data
    hide_data = ctk.CTkButton(btn_frame2, text='Hide Data', 
                        width=200, height=40,
                        font=('Arial', 18),
                        corner_radius=10,
                        bg_color='#153c4a', 
                        hover_color='blue',
                        command=hidedata
                        )
    hide_data.place(x=10, y=10)


    # Show data
    show_data = ctk.CTkButton(btn_frame2, text='Show Data', 
                        width=200, height=40,
                        font=('Arial', 18),
                        corner_radius=10,
                        bg_color='#153c4a', 
                        hover_color='blue',
                        command=showdata
                        )
    show_data.place(x=220, y=10)



    root.mainloop()


    
    
    
    ################### hide data ends ##########################
    



def open_ide():
    root = Toplevel()
    root.title("*Untitled  -  Tiny-iDE")
    root.geometry("1280x770+150+50")
    root.configure(background='grey')
    root.resizable(False, False)
    
    
    ######################## IDE Starts #####################################


    # code area
    code_area = ctk.CTkTextbox(root, font=(
        'consolas', 18), undo=True, width=1155, height=600)
    code_area.place(x=120, y=0)

    # output
    show_output = ctk.CTkTextbox(root, font=('consolas', 12),
                                fg_color='black', text_color='lightgreen',
                                width=1155, height=170)
    show_output.place(x=120, y=600)


    # menu commands
    def new_file(*args):
        root.title("*newfile  -  Tiny-iDE")
        code_area.delete(1.0, END)


    def python_file(*args):
        save_file()


    def open_file(*args):
        global file
        root.lift()
        file = askopenfilename(defaultextension="*.py",
                            filetypes=[
                                ("Python Files", "*.py"),
                                ("Text Document", "*.txt"),
                                ("Other Files", "*.*")
                            ], parent=root)

        with open(file, 'r') as f:
            content = f.read()
            code_area.delete(1.0, END)
            code_area.insert(1.0, content)

            root.title(os.path.basename(file) + "  -  Tiny-iDE")


    def save_file(*args):
        global file
        root.lift()
        if file == None:
            file = asksaveasfilename(initialfile="newfile.py",
                                    defaultextension='.py',
                                    filetypes=[
                                        ("Python File", "*.py"),
                                        ("Text Document", "*.txt"),
                                        ("Other Files", "*.*")
                                    ], parent=root)

            if file == "":
                file = None
            else:
                f = open(file, 'w')
                f.write(code_area.get(1.0, END))
                f.close()
                root.title(os.path.basename(file) + "  -  Tiny-iDE")

        else:
            f = open(file, 'w')
            f.write(code_area.get(1.0, END))
            f.close()
            root.title(os.path.basename(file) + "  -  Tiny-iDE")


    def save_as(*args):
        global file
        root.lift()
        file = asksaveasfilename(defaultextension=".py",
                                initialfile="newfile.py",
                                filetypes=[
                                    ("python File", "*.py"),
                                    ("Text Document", "*.txt"),
                                    ("Other Files", "*.*")
                                ], parent=root)

        with open(file, "w") as f:
            content = code_area.get(1.0, END)
            f.write(content)

            root.title(os.path.basename(file) + "  -  Tiny-iDE")


    # work of commands
    def select_language(*args):
        root.lift()
        messagebox.showinfo(
            "Language", "Python was found! cannot find other languages!", parent=root)


    def run_file(*args):
        global file

        if file is None:
            root.lift()
            save = messagebox.askquestion(
                "Save", "You need to save first. Do you want to save the file?", parent=root)
            if save == "yes":
                save_file()
            else:
                return  # If the user chooses not to save, return without running the file

        command = f"python3 {file}"  # Add "python3" before the file path
        process = subprocess.Popen(
            command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
        output, error = process.communicate()
        # Use END instead of 1.0 to insert at the end of the text widget
        show_output.insert(END, output)
        show_output.insert(END, error)


    def setting(*args):
        s1 = settings.Settings()


    def python(*args):
        pass


    def user_id(*args):
        signup = SignUp()


    def github_profile(*args):
        webbrowser.open('github.com')


    def information(*args):
        root.lift()
        messagebox.showinfo(
            "Information", "This is liteweight IDE, made with python(Tkinter).", parent=root)


    def exit_window(*args):
        root.destroy()


    def undo(*args):
        try:
            code_area.edit_undo()
        except:
            pass


    def redo(*args):
        try:
            code_area.edit_redo()
        except:
            pass


    def cut(*args):
        code_area.event_generate('<<Cut>>')


    def copy(*args):
        code_area.event_generate('<<Copy>>')


    def paste(*args):
        code_area.event_generate('<<Paste>>')


    def expand_output(*args):
        pass


    def max_window(*args):
        pass


    def how_work():
        root.lift()
        messagebox.showwarning('User Manual', 'Please check the user manual pdf.', parent=root)


    def all_command(*args):
        key = keyboard_shortcut.KeyboardShortcuts()


    def source_code():
        webbrowser.open('https://github.com/mdshakib007/Tkinter-Projects')


    def report():
        r1 = report_page.Report()


    def about(*args):
        root.lift()
        messagebox.showinfo('About', 'This is most liteweight project for making code is fun!', parent=root)
        
        
        
        
    # bind keyboard shortcuts
    root.bind('<Control-n>', new_file)
    root.bind('<Control-p>', python_file)
    root.bind('<Control-o>', open_file)
    root.bind('<Control-s>', save_file)
    root.bind('<Control-q>', exit_window)
    root.bind('<Control-z>', undo)
    root.bind('<Control-y>', redo)
    root.bind('<Control-x>', cut)
    root.bind('<Control-c>', copy)
    root.bind('<Control-v>', paste)
    root.bind('<Control-r>', run_file)
    root.bind('<Control-e>', expand_output)
    root.bind('<Alt-m>', max_window)
    root.bind('<Alt-s>', source_code)
    root.bind('<Alt-a>', about)


    # Menu
    ### File menu ###
    main_menu = Menu(root)

    file_menu = Menu(main_menu, tearoff=0, font='Arial 14')

    file_menu.add_command(label="New File", accelerator="Ctrl+N", command=new_file)
    file_menu.add_command(label="Python File",
                        accelerator="Ctrl+P", command=python_file)
    file_menu.add_command(
        label="Open file", accelerator="Ctrl+O", command=open_file)
    file_menu.add_separator()
    file_menu.add_command(label="Save", accelerator="Ctrl+S", command=save_file)
    file_menu.add_command(label="Save as...",
                        accelerator="Ctrl+S", command=save_as)
    file_menu.add_separator()
    file_menu.add_command(label="Exit", accelerator="Ctrl+Q", command=exit_window)

    main_menu.add_cascade(menu=file_menu, label="File", font="Arial 13")


    # Edit Menu
    edit_menu = Menu(main_menu, tearoff=0, font='Arial 14')

    edit_menu.add_command(label="Undo", accelerator='Ctrl+Z', command=undo)
    edit_menu.add_command(label="Redo", accelerator='Ctrl+Y', command=redo)
    edit_menu.add_separator()
    edit_menu.add_command(label="Cut", accelerator='Ctrl+X', command=cut)
    edit_menu.add_command(label="Copy", accelerator='Ctrl+C', command=copy)
    edit_menu.add_command(label="Paste", accelerator='Ctrl+V', command=paste)
    edit_menu.add_separator()
    edit_menu.add_command(label="Expand Output Window",
                        accelerator='Ctrl+E', command=expand_output)
    edit_menu.add_command(label="Max Window",
                        accelerator='Alt+M', command=max_window)
    edit_menu.add_command(label="Exit", accelerator='Ctrl+Q', command=exit_window)

    main_menu.add_cascade(menu=edit_menu, label='Edit', font='Arial 13')


    # Run Menu
    run_menu = Menu(main_menu, tearoff=0, font='Arial 14')

    run_menu.add_command(label='Run',
                        accelerator='Ctrl+R', command=run_file)
    run_menu.add_command(label='Run And Debug',
                        accelerator='Ctrl+R+D', command=run_file)
    run_menu.add_separator()
    run_menu.add_command(
        label='Settings', accelerator='Ctrl+Alt+S', command=setting)
    run_menu.add_command(label='Quit', accelerator='Ctrl+Q', command=exit_window)

    main_menu.add_cascade(menu=run_menu, label='Run', font='Arial 13')


    # Help menu
    help_menu = Menu(main_menu, tearoff=0, font='Arial 14')

    help_menu.add_command(label="How It's Work?", command=how_work)
    help_menu.add_command(label="Show All Commands...",
                        accelerator='Ctrl+C+A', command=all_command)
    help_menu.add_command(label="Get Source Code", command=source_code)
    help_menu.add_separator()
    help_menu.add_command(label="Report...", command=report)
    help_menu.add_command(label="About", command=about)

    main_menu.add_cascade(menu=help_menu, label="Help", font='Arial 13')

    root.config(menu=main_menu)


    ### buttons ###
    # select language
    select = Image.open("assets/ide-language.png")
    s_res = select.resize((30, 30))
    se_lang = ImageTk.PhotoImage(image=s_res)
    se_btn = ctk.CTkButton(root, image=se_lang,
                        bg_color='grey',
                        command=select_language, height=30, width=30,
                        text='',
                        fg_color='grey',
                        hover_color='white',
                        
                        )
    se_btn.place(x=25, y=100)

    # run
    run_ = Image.open("assets/ide-run.png")
    run_res = run_.resize((30, 30))
    run = ImageTk.PhotoImage(image=run_res)
    run_btn = ctk.CTkButton(root, image=run, bg_color='grey', command=run_file, height=30, width=30,
                            text='',
                            fg_color='grey',
                            hover_color='white')
    run_btn.place(x=25, y=170)
    

    # settings
    setting_ = Image.open("assets/ide-settings.png")
    setting_res = setting_.resize((30, 30))
    setting_img = ImageTk.PhotoImage(image=setting_res)
    setting_btn = ctk.CTkButton(root, image=setting_img, bg_color='grey',
                                command=setting, height=30, width=30,
                                text='',
                                fg_color='grey',
                                hover_color='white')
    setting_btn.place(x=25, y=240)
    

    # python
    py_ = Image.open("assets/python.png")
    py_res = py_.resize((30, 30))
    py = ImageTk.PhotoImage(image=py_res)
    py_btn = ctk.CTkButton(root, image=py, bg_color='grey', command=python, height=30, width=30,
                        text='',
                        fg_color='grey',
                        hover_color='white')
    py_btn.place(x=25, y=310)


    # user
    user_ = Image.open("assets/user2.png")
    user_res = user_.resize((30, 30))
    user = ImageTk.PhotoImage(image=user_res)
    user_btn = ctk.CTkButton(root, image=user, bg_color='grey', command=user_id,
                            height=30, width=30,
                            text='',
                            fg_color='grey',
                            hover_color='white')
    user_btn.place(x=25, y=380)


    # github
    github_ = Image.open("assets/git.png")
    github_res = github_.resize((30, 30))
    github = ImageTk.PhotoImage(image=github_res)
    github_btn = ctk.CTkButton(root, image=github, bg_color='grey',
                            command=github_profile,
                            height=30, width=30,
                            text='',
                            fg_color='grey',
                            hover_color='white')
    github_btn.place(x=25, y=460)
    

    # info
    info_ = Image.open("assets/ide-info.png")
    info_res = info_.resize((30, 30))
    info = ImageTk.PhotoImage(image=info_res)
    info_btn = ctk.CTkButton(root, image=info, bg_color='grey',
                            command=information,
                            height=30, width=30,
                            text='',
                            fg_color='grey',
                            hover_color='white')
    info_btn.place(x=25, y=530)
    

    # cancel
    cancel_ = Image.open("assets/cancel.png")
    cancel_res = cancel_.resize((30, 30))
    cancel = ImageTk.PhotoImage(image=cancel_res)
    cancel_btn = ctk.CTkButton(root, image=cancel,
                            bg_color='grey',
                            command=exit_window,
                            height=30, width=30,
                            text='',
                            fg_color='grey',
                            hover_color='white')
    cancel_btn.place(x=25, y=600)


    
    ########################### IDE Ends ###################################
    


# for quick exit
root.bind('<Escape>', exit)


# Main frame
mainFrame = ctk.CTkFrame(root, 
                         width=650, 
                         height=800,
                         fg_color='#d4c1e6',
                         border_width=2,
                         corner_radius=10,)
mainFrame.place(x=10, y=10)

uf = ctk.CTkFrame(mainFrame,
                  width=200,
                  height=40,
                  fg_color='#d4c1e6',
                  border_width=1,
                  corner_radius=10
                  )
uf.place(x=250, y=10)


# label for plot username
show_username = Label(uf, 
             text=username,
             bg='#d4c1e6',
             font='Arial 14 bold'
             )

show_username.place(x=5, y=7)




# Create button for studentRegi.png
studentRegi_image = Image.open('assets/studentregi.png')
studentRegi_image_resize = studentRegi_image.resize((200, 200))
studentRegi_photo = ImageTk.PhotoImage(image=studentRegi_image_resize)

studentRegi_button = ctk.CTkButton(mainFrame, 
                                width=200, 
                                height=200,
                                image=studentRegi_photo,
                                fg_color='#d4c1e6',
                                hover=False,
                                cursor=('hand2'),
                                text='',
                                command=student_registration)
studentRegi_button.image = studentRegi_photo
studentRegi_button.place(x=5, y=5)


# Create button for add_user.png
add_user_image = Image.open('assets/add_user.png')
add_user_image_resize = add_user_image.resize((50,50))
add_user_photo = ImageTk.PhotoImage(image=add_user_image_resize)

add_user_button = ctk.CTkButton(mainFrame, 
                                width=80, 
                                height=80,
                                image=add_user_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text=' Sign Up',
                                compound='top',
                                text_color='black',
                                font=('Arial', 15, 'bold'),
                                command=singup_form
                                )
add_user_button.image = add_user_photo
add_user_button.place(x=560, y=5)




### label 
Label(mainFrame, text='| Services:',
      font='Arial 13',
      bd=0, 
      bg='#d4c1e6',
      fg='grey').place(x=30, y=260)




# art and design button
art_image = Image.open('assets/draw2.png')
art_image_resize = art_image.resize((100, 100))
art_photo = ImageTk.PhotoImage(image=art_image_resize)

art_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=art_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='',
                                command=open_writingboard)
art_button.image = art_photo
art_button.place(x=20, y=300)

# download button
download_image = Image.open('assets/download.png')
download_image_resize = download_image.resize((90, 90))
download_photo = ImageTk.PhotoImage(image=download_image_resize)

download_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=download_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='',
                                command=open_downloadar)

download_button.image = download_photo
download_button.place(x=150, y=300)

# calculator button
calculator_image = Image.open('assets/math.png')
calculator_image_resize = calculator_image.resize((100, 100))
calculator_photo = ImageTk.PhotoImage(image=calculator_image_resize)

calculator_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=calculator_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='',
                                command=open_calculator)

calculator_button.image = calculator_photo
calculator_button.place(x=280, y=300)

# notepad button
notepad_image = Image.open('assets/notepad.png')
notepad_image_resize = notepad_image.resize((100, 100))
notepad_photo = ImageTk.PhotoImage(image=notepad_image_resize)

notepad_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=notepad_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='',
                                command=open_notepad)

notepad_button.image = notepad_photo
notepad_button.place(x=400, y=300)

# ide button
ide_image = Image.open('assets/ide.png')
ide_image_resize = ide_image.resize((100, 100))
ide_photo = ImageTk.PhotoImage(image=ide_image_resize)

ide_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=ide_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='',
                                command=open_ide)
ide_button.image = ide_photo
ide_button.place(x=520, y=300)

# translate button
translate_image = Image.open('assets/translate.png')
translate_image_resize = translate_image.resize((90, 90))
translate_photo = ImageTk.PhotoImage(image=translate_image_resize)

translate_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=translate_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='', 
                                command=translate_anything)

translate_button.image = translate_photo
translate_button.place(x=20, y=420)


# apartment button
apartment_image = Image.open('assets/apartment.png')
apartment_image_resize = apartment_image.resize((90, 90))
apartment_photo = ImageTk.PhotoImage(image=apartment_image_resize)

apartment_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=apartment_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='', 
                                command=house_price)

apartment_button.image = apartment_photo
apartment_button.place(x=150, y=420)


# calendar button
calendar_image = Image.open('assets/clock.png')
calendar_image_resize = calendar_image.resize((100, 100))
calendar_photo = ImageTk.PhotoImage(image=calendar_image_resize)

calendar_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=calendar_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='', 
                                command=open_calendar)

calendar_button.image = calendar_photo
calendar_button.place(x=280, y=420)


# dictionary button
dictionary_image = Image.open('assets/dictionary.png')
dictionary_image_resize = dictionary_image.resize((80, 80))
dictionary_photo = ImageTk.PhotoImage(image=dictionary_image_resize)

dictionary_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=dictionary_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='', 
                                command=open_dictionary)

dictionary_button.image = dictionary_photo
dictionary_button.place(x=400, y=420)


# grade button
grade_image = Image.open('assets/calculator.png')
grade_image_resize = grade_image.resize((90, 90))
grade_photo = ImageTk.PhotoImage(image=grade_image_resize)

grade_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=grade_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='', 
                                command=grade_calc)

grade_button.image = grade_photo
grade_button.place(x=520, y=420)



# hide data in picture button
hide_image = Image.open('assets/encryption.png')
hide_image_resize = hide_image.resize((80, 80))
hide_photo = ImageTk.PhotoImage(image=hide_image_resize)

hide_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=hide_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='', 
                                command=data_hide)

hide_button.image = hide_photo
hide_button.place(x=20, y=540)



# encryption decryption button
encrypt_image = Image.open('assets/encrypt1.png')
encrypt_image_resize = encrypt_image.resize((100, 100))
encrypt_photo = ImageTk.PhotoImage(image=encrypt_image_resize)

encrypt_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=encrypt_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='', 
                                command=data_encryption)

encrypt_button.image = encrypt_photo
encrypt_button.place(x=150, y=540)


# excel button
excel_image = Image.open('assets/xlsx.png')
excel_image_resize = excel_image.resize((80, 80))
excel_photo = ImageTk.PhotoImage(image=excel_image_resize)

excel_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=excel_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='', 
                                command=excel_file)

excel_button.image = encrypt_photo
excel_button.place(x=280, y=540)




####### games ##########

Label(mainFrame, text='| Games:', font='Arial 13', 
      fg='grey', bg='#d4c1e6').place(x=20, y=650)


# quiz game button
quiz_image = Image.open('assets/quiz2.png')
quiz_image_resize = quiz_image.resize((80, 80))
quiz_photo = ImageTk.PhotoImage(image=quiz_image_resize)

quiz_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=quiz_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='',
                                command=quiz_game)
quiz_button.image = quiz_photo
quiz_button.place(x=20, y=680)


# snake game button
snake_image = Image.open('assets/snake.png')
snake_image_resize = snake_image.resize((100, 100))
snake_photo = ImageTk.PhotoImage(image=snake_image_resize)

snake_button = ctk.CTkButton(mainFrame, 
                                width=100, 
                                height=100,
                                image=snake_photo,
                                fg_color='#d4c1e6',
                                hover_color='#a77af5', 
                                text='')
snake_button.image = snake_photo
snake_button.place(x=150, y=680)



### sidebar ### 

## frame
sideFrame = ctk.CTkFrame(root, 
                         width=100, 
                         height=600,
                         fg_color='#d4c1e6',
                         border_width=1, 
                         corner_radius=10,)
sideFrame.place(x=700, y=50)

# label 
Label(sideFrame, text='Quick Links:',
      font='Arial 10 bold',
      bd=0, 
      bg='#d4c1e6',
      fg='grey').place(x=5, y=5)



# account button
account_image = Image.open('assets/user2.png')
account_image_resize = account_image.resize((50, 50))
account_photo = ImageTk.PhotoImage(image=account_image_resize)

account_button = ctk.CTkButton(sideFrame, 
                                width=50, 
                                height=50,
                                image=account_photo,
                                fg_color='#d4c1e6',
                                hover_color='#97939e', 
                                text='',
                                command=open_account)
account_button.image = account_photo
account_button.place(x=15, y=40)


# chatgpt button
chatgpt_image = Image.open('assets/chat.png')
chatgpt_image_resize = chatgpt_image.resize((50, 50))
chatgpt_photo = ImageTk.PhotoImage(image=chatgpt_image_resize)

chatgpt_button = ctk.CTkButton(sideFrame, 
                                width=50, 
                                height=50,
                                image=chatgpt_photo,
                                fg_color='#d4c1e6',
                                hover_color='grey', 
                                text='',
                                command=open_chatgpt)
chatgpt_button.image = chatgpt_photo
chatgpt_button.place(x=15, y=100)


# github button
github_image = Image.open('assets/github.png')
github_image_resize = github_image.resize((50, 50))
github_photo = ImageTk.PhotoImage(image=github_image_resize)

github_button = ctk.CTkButton(sideFrame, 
                                width=50, 
                                height=50,
                                image=github_photo,
                                fg_color='#d4c1e6',
                                hover_color='grey', 
                                text='',
                                command=open_github)
github_button.image = github_photo
github_button.place(x=15, y=160)


# stackoverflow button
sof_image = Image.open('assets/sof.png')
sof_image_resize = sof_image.resize((50, 50))
sof_photo = ImageTk.PhotoImage(image=sof_image_resize)

sof_button = ctk.CTkButton(sideFrame, 
                                width=50, 
                                height=50,
                                image=sof_photo,
                                fg_color='#d4c1e6',
                                hover_color='grey', 
                                text='',
                                command=open_sof)
sof_button.image = sof_photo
sof_button.place(x=15, y=220)


# python button
python_image = Image.open('assets/python.png')
python_image_resize = python_image.resize((50, 50))
python_photo = ImageTk.PhotoImage(image=python_image_resize)

python_button = ctk.CTkButton(sideFrame, 
                                width=50, 
                                height=50,
                                image=python_photo,
                                fg_color='#d4c1e6',
                                hover_color='grey', 
                                text='',
                                command=open_python)
python_button.image = python_photo
python_button.place(x=15, y=280)


# facebook button
fb_image = Image.open('assets/facebook.png')
fb_image_resize = fb_image.resize((50, 50))
fb_photo = ImageTk.PhotoImage(image=fb_image_resize)

fb_button = ctk.CTkButton(sideFrame, 
                                width=50, 
                                height=50,
                                image=fb_photo,
                                fg_color='#d4c1e6',
                                hover_color='grey', 
                                text='',
                                command=open_facebook)
fb_button.image = fb_photo
fb_button.place(x=15, y=340)


# youtube button
yt_image = Image.open('assets/youtube.png')
yt_image_resize = yt_image.resize((50, 50))
yt_photo = ImageTk.PhotoImage(image=yt_image_resize)

yt_button = ctk.CTkButton(sideFrame, 
                                width=50, 
                                height=50,
                                image=yt_photo,
                                fg_color='#d4c1e6',
                                hover_color='grey', 
                                text='',
                                command=open_youtube)
yt_button.image = yt_photo
yt_button.place(x=15, y=400)


# share button
share_image = Image.open('assets/share.png')
share_image_resize = share_image.resize((50, 50))
share_photo = ImageTk.PhotoImage(image=share_image_resize)

share_button = ctk.CTkButton(sideFrame, 
                                width=50, 
                                height=50,
                                image=share_photo,
                                fg_color='#d4c1e6',
                                hover_color='grey', 
                                text='',
                                command=share)
share_button.image = yt_photo
share_button.place(x=15, y=460)


# info button
info_image = Image.open('assets/info2.png')
info_image_resize = info_image.resize((50, 50))
info_photo = ImageTk.PhotoImage(image=info_image_resize)

info_button = ctk.CTkButton(sideFrame, 
                                width=50, 
                                height=50,
                                image=info_photo,
                                fg_color='#d4c1e6',
                                hover_color='grey', 
                                text='',
                                command=information)
info_button.image = info_photo
info_button.place(x=15, y=400)


# bangladesh button
bd_image = Image.open('assets/bd.png')
bd_image_resize = bd_image.resize((50, 50))
bd_photo = ImageTk.PhotoImage(image=bd_image_resize)

bd_button = ctk.CTkButton(sideFrame, 
                                width=50, 
                                height=50,
                                image=bd_photo,
                                fg_color='#d4c1e6',
                                hover_color='grey', 
                                text='',
                                command=bangladesh)
bd_button.image = bd_photo
bd_button.place(x=15, y=460)


# report button
report_image = Image.open('assets/report.png')
report_image_resize = report_image.resize((50, 50))
report_photo = ImageTk.PhotoImage(image=report_image_resize)

report_button = ctk.CTkButton(sideFrame, 
                                width=50, 
                                height=50,
                                image=report_photo,
                                fg_color='#d4c1e6',
                                hover_color='grey', 
                                text='',
                                command=report_issue)
report_button.image = report_photo
report_button.place(x=15, y=520)




root.mainloop()
